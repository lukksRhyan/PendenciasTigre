import json
from openpyxl import Workbook
from openpyxl.styles.builtins import output
from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox


class XlsxHandler:
    def __init__(self,json_file_path):
        self.json_file_path = json_file_path

    def json_to_excel(self):
        # Carregar o arquivo JSON
        with open(self.json_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)

        # Criar um novo workbook do Excel
        wb = Workbook()

        # ===== PLANILHA NOTA MÃE =====
        if 'nºnotamãe' in data:
            nota_mae = data['nºnotamãe']
            ws_mae = wb.active
            ws_mae.title = "Nota Mãe"

            # Cabeçalho da nota mãe
            ws_mae.append(["Número da Nota", nota_mae["Número da Nota"]])
            ws_mae.append(["CFOP", nota_mae["CFOP"]])
            ws_mae.append(["Total", nota_mae["Total"]])
            #ws_mae.append(["Informações Adicionais", nota_mae["Informações Adicionais"]])
            ws_mae.append([])  # Linha em branco

            # Cabeçalho dos produtos
            ws_mae.append(["Código", "Descrição", "Quantidade", "Unidade"])

            # Adicionar produtos
            for produto in nota_mae["Produtos"]:
                ws_mae.append([
                    produto["Código"].lstrip("0"),
                    produto["Descrição"],
                    produto["Quantidade"],
                    produto["Unidade"]
                ])

            # Ajustar largura das colunas
            for col in ws_mae.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_mae.column_dimensions[column].width = adjusted_width

        # ===== PLANILHA NOTAS FILHAS =====
        if 'notas_filhas' in data and data['notas_filhas']:
            for i, nota_filha in enumerate(data['notas_filhas'], 1):
                ws_filha = wb.create_sheet(title=f"Nota Filha {i}")

                # Cabeçalho da nota filha
                ws_filha.append(["Número da Nota", nota_filha["Número da Nota"]])
                ws_filha.append(["CFOP", nota_filha["CFOP"]])
                ws_filha.append(["Total", nota_filha["Total"]])
                #ws_filha.append(["Informações Adicionais", nota_filha["Informações Adicionais"]])
                ws_filha.append([])  # Linha em branco

                # Cabeçalho dos produtos
                ws_filha.append(["Código", "Descrição", "Quantidade", "Unidade"])

                # Adicionar produtos
                for produto in nota_filha["Produtos"]:
                    ws_filha.append([
                        produto["Código"].lstrip('0'),
                        produto["Descrição"],
                        produto["Quantidade"],
                        produto["Unidade"]
                    ])

                # Ajustar largura das colunas
                for col in ws_filha.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws_filha.column_dimensions[column].width = adjusted_width

        # ===== PLANILHA PRODUTOS RESTANTES =====
        if 'produtos_restantes' in data and data['produtos_restantes']:
            ws_restantes = wb.create_sheet(title="Produtos Restantes")

            # Cabeçalho
            ws_restantes.append(["Código", "Descrição", "Quantidade", "Unidade"])

            # Adicionar produtos restantes
            for produto in data['produtos_restantes']:
                ws_restantes.append([
                    produto["Código"].lstrip('0'),
                    produto["Descrição"],
                    produto["Quantidade"],
                    produto["Unidade"]
                ])

            # Ajustar largura das colunas
            for col in ws_restantes.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_restantes.column_dimensions[column].width = adjusted_width

        output_excel_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
            title="Salvar como"
        )

        # Se o usuário cancelar, não salvar
        if not output_excel_path:
            print("Operação cancelada pelo usuário.")
            return
        try:
            wb.save(output_excel_path)
        except Exception as error:
            messagebox.showwarning("Aviso", "❌ Erro ao salvar o arquivo. \nVerifique se não está aberto em outro programa")
            print(error)
        finally:
            messagebox.showinfo("Sucesso", f"✅ Arquivo {output_excel_path} criado com sucesso!")
        print(f"Arquivo Excel gerado com sucesso: {output_excel_path}")